from datetime import datetime, timedelta
from django.db.models import Count, Sum, Q
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import (
    TemplateView, ListView, CreateView, UpdateView, DetailView, DeleteView
)
from .models import Cliente, Servicio, HistorialEstado, EstadoServicio
from .forms import ClienteForm, ServicioForm, HistorialEstadoForm
from .models import Servicio
from .models import Cliente
from io import BytesIO
from datetime import datetime, time

from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Sum
from django.db.models.functions import Coalesce

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from decimal import Decimal
from django.db.models import Sum

from decimal import Decimal
from django.db.models import Count
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)

from django.views.generic import DetailView





# Home: resumen de estados
class HomeView(TemplateView):
    template_name = "core/index.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        counts = dict(
            Servicio.objects.values_list("estado_actual")
            .annotate(c=Count("id"))
            .values_list("estado_actual", "c")
        )

        resumen = [
            {"estado": EstadoServicio.PENDIENTE, "count": counts.get(EstadoServicio.PENDIENTE, 0), "clase": "secondary", "icono": "bi-clock"},
            {"estado": EstadoServicio.DIAGNOSTICO, "count": counts.get(EstadoServicio.DIAGNOSTICO, 0), "clase": "info", "icono": "bi-search"},
            {"estado": EstadoServicio.REPARACION, "count": counts.get(EstadoServicio.REPARACION, 0), "clase": "primary", "icono": "bi-wrench"},
            {"estado": EstadoServicio.ESPERA_REPUESTOS, "count": counts.get(EstadoServicio.ESPERA_REPUESTOS, 0), "clase": "warning", "icono": "bi-hourglass-split"},
            {"estado": EstadoServicio.LISTO_RETIRO, "count": counts.get(EstadoServicio.LISTO_RETIRO, 0), "clase": "success", "icono": "bi-check-circle"},
        ]
        ctx["servicios_resumen"] = resumen
        return ctx

# ---- Clientes ----
from django.views.generic import ListView
from django.db.models import Q, Count, Max
from .models import Cliente

class ClienteListView(ListView):
    model = Cliente
    template_name = "core/clientes_list.html"
    paginate_by = 20

    def get_queryset(self):
        qs = (
            super()
            .get_queryset()
            .annotate(
                servicios_count=Count("servicios", distinct=True),
                ultimo_ingreso=Max("servicios__fecha_ingreso"),
            )
            .order_by("-ultimo_ingreso", "apellido", "nombre")

        )

        q = (self.request.GET.get("q") or "").strip()
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q) |
                Q(apellido__icontains=q) |
                Q(telefono__icontains=q) |
                Q(email__icontains=q) |
                Q(dni_cuit__icontains=q) |
                Q(direccion__icontains=q)
            )
        return qs


class ClienteCreateView(CreateView):
    model = Cliente
    form_class = ClienteForm
    template_name = "core/clientes_form.html"
    success_url = reverse_lazy("clientes_list")

class ClienteUpdateView(UpdateView):
    model = Cliente
    form_class = ClienteForm
    template_name = "core/clientes_form.html"
    success_url = reverse_lazy("clientes_list")

class ClienteDeleteView(DeleteView):
    model = Cliente
    template_name = "core/confirm_delete.html"
    success_url = reverse_lazy("clientes_list")

# ---- Servicios ----
from django.views.generic import ListView
from django.db.models import Q, Count
from .models import Servicio

class ServicioListView(ListView):
    model = Servicio
    template_name = "core/servicios_list.html"
    paginate_by = 20

    def get_queryset(self):
        qs = (
            super()
            .get_queryset()
            .select_related("cliente")
            .order_by("-fecha_ingreso")
        )

        q = (self.request.GET.get("q") or "").strip()
        cliente_id = (self.request.GET.get("cliente") or "").strip()
        estado = (self.request.GET.get("estado") or "").strip()

        if cliente_id:
            qs = qs.filter(cliente_id=cliente_id)

        if q:
            qs = qs.filter(
                Q(id__icontains=q)
                | Q(cliente__nombre__icontains=q)
                | Q(cliente__apellido__icontains=q)
                | Q(cliente__dni_cuit__icontains=q)
                | Q(cliente__telefono__icontains=q)
                | Q(tipo_equipo__icontains=q)
                | Q(marca_modelo__icontains=q)
            )

        # Guardamos el queryset "base" (sin estado) para contar por estados
        self.base_qs = qs

        # Aplicamos filtro por estado si viene
        if estado:
            qs = qs.filter(estado_actual=estado)

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        ctx.update(_build_report_data(self.request))


        # ✅ Conteos por estado (respetan q + cliente, pero NO el estado seleccionado)
        base_counts = dict(
            self.base_qs.values_list("estado_actual")
            .annotate(c=Count("id"))
            .values_list("estado_actual", "c")
        )
        total = sum(base_counts.values())

        quick_filters = [
            {"value": "Pendiente", "label": "Pendiente", "btn": "secondary", "icon": "bi-clock"},
            {"value": "En Diagnóstico", "label": "Diagnóstico", "btn": "info", "icon": "bi-search"},
            {"value": "En Reparación", "label": "Reparación", "btn": "primary", "icon": "bi-wrench"},
            {"value": "En Espera de Repuestos", "label": "Repuestos", "btn": "warning", "icon": "bi-hourglass-split"},
            {"value": "Listo para Retirar", "label": "Listo", "btn": "success", "icon": "bi-check-circle"},
            {"value": "Entregado", "label": "Entregado", "btn": "dark", "icon": "bi-check2-all"},
            {"value": "Cancelado", "label": "Cancelado", "btn": "danger", "icon": "bi-x-circle"},
        ]

        for f in quick_filters:
            f["count"] = base_counts.get(f["value"], 0)

        ctx["quick_filters"] = quick_filters
        ctx["quick_total"] = total

        return ctx

class ServicioCreateView(CreateView):
    model = Servicio
    form_class = ServicioForm
    template_name = "core/servicios_form.html"

    def form_valid(self, form):
        resp = super().form_valid(form)
        HistorialEstado.objects.create(
            servicio=self.object,
            estado=self.object.estado_actual,
            observaciones="Estado inicial al crear el servicio",
        )
        return resp

    def get_success_url(self):
        return reverse_lazy("servicio_detalle", kwargs={"pk": self.object.pk})

class ServicioUpdateView(UpdateView):
    model = Servicio
    form_class = ServicioForm
    template_name = "core/servicios_form.html"

    def form_valid(self, form):
        # si cambió estado, guardamos en historial
        old = Servicio.objects.get(pk=self.object.pk)
        resp = super().form_valid(form)
        if old.estado_actual != self.object.estado_actual:
            HistorialEstado.objects.create(
                servicio=self.object,
                estado=self.object.estado_actual,
                observaciones=self.request.POST.get("observaciones_historial") or "Cambio de estado",
            )
        return resp

    def get_success_url(self):
        return reverse_lazy("servicio_detalle", kwargs={"pk": self.object.pk})

class ServicioDetailView(DetailView):
    model = Servicio
    template_name = "core/servicio_detalle.html"

class ServicioDeleteView(DeleteView):
    model = Servicio
    template_name = "core/confirm_delete.html"
    success_url = reverse_lazy("servicios_list")

class ServicioPrintView(DetailView):
    model = Servicio
    template_name = "core/servicio_imprimir.html"

# ---- Reportes ----

class ReportesView(TemplateView):
    template_name = "core/reportes.html"

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        fecha_inicio_str = self.request.GET.get("fecha_inicio") or ""
        fecha_fin_str = self.request.GET.get("fecha_fin") or ""
        estado = (self.request.GET.get("estado") or "").strip()  # ✅ NUEVO

        fecha_inicio = None
        fecha_fin = None
        try:
            if fecha_inicio_str:
                fecha_inicio = datetime.strptime(fecha_inicio_str, "%Y-%m-%d").date()
            if fecha_fin_str:
                fecha_fin = datetime.strptime(fecha_fin_str, "%Y-%m-%d").date()
        except ValueError:
            fecha_inicio = None
            fecha_fin = None

        # Ingresos: por fecha_entrega (solo servicios entregados)
        ingresos_qs = Servicio.objects.filter(fecha_entrega__isnull=False)
        if fecha_inicio:
            ingresos_qs = ingresos_qs.filter(fecha_entrega__date__gte=fecha_inicio)
        if fecha_fin:
            ingresos_qs = ingresos_qs.filter(fecha_entrega__date__lte=fecha_fin)

        # ✅ Si filtran por estado, aplicalo también a ingresos
        if estado:
            ingresos_qs = ingresos_qs.filter(estado_actual=estado)

        total_ingresos = ingresos_qs.aggregate(total=Sum("costo_total"))["total"] or 0

        # Mes actual
        hoy = timezone.localdate()
        inicio_mes = hoy.replace(day=1)
        next_month = (inicio_mes + timedelta(days=32)).replace(day=1)
        fin_mes = next_month - timedelta(days=1)

        ingresos_mes_actual = ingresos_qs.filter(
            fecha_entrega__date__gte=inicio_mes,
            fecha_entrega__date__lte=fin_mes,
        ).aggregate(total=Sum("costo_total"))["total"] or 0

        # Servicios: por fecha_ingreso
        servicios_qs = Servicio.objects.all()
        if fecha_inicio:
            servicios_qs = servicios_qs.filter(fecha_ingreso__date__gte=fecha_inicio)
        if fecha_fin:
            servicios_qs = servicios_qs.filter(fecha_ingreso__date__lte=fecha_fin)

        # ✅ Filtro por estado en servicios
        if estado:
            servicios_qs = servicios_qs.filter(estado_actual=estado)

        servicios_por_estado = list(
            servicios_qs.values("estado_actual")
            .annotate(count=Count("id"))
            .order_by("-count")
        )
        total_servicios_filtrados = sum(item["count"] for item in servicios_por_estado)

        servicios_por_tipo_equipo = list(
            servicios_qs.values("tipo_equipo")
            .annotate(count=Count("id"))
            .order_by("-count")
        )

        servicios_por_cliente = list(
            servicios_qs.values("cliente__apellido", "cliente__nombre", "cliente__id")
            .annotate(count=Count("id"))
            .order_by("-count")[:10]
        )

        # ✅ Choices para llenar el <select>
        estado_choices = Servicio._meta.get_field("estado_actual").choices

        ctx.update({
            "fecha_inicio_str": fecha_inicio_str,
            "fecha_fin_str": fecha_fin_str,
            "estado_actual": estado,          # ✅ para marcar selected
            "estado_choices": estado_choices, # ✅ para poblar el select

            "ingresos_mes_actual": ingresos_mes_actual,
            "total_ingresos": total_ingresos,
            "servicios_por_estado": servicios_por_estado,
            "total_servicios_filtrados": total_servicios_filtrados,
            "servicios_por_tipo_equipo": servicios_por_tipo_equipo,
            "servicios_por_cliente": servicios_por_cliente,
        })
        return ctx


def _parse_date_range(request):
    """
    Lee ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD (opcionales).
    Devuelve (dt_desde, dt_hasta) en timezone local.
    """
    tz = timezone.get_current_timezone()
    desde_s = (request.GET.get("desde") or "").strip()
    hasta_s = (request.GET.get("hasta") or "").strip()

    dt_desde = None
    dt_hasta = None

    if desde_s:
        d = datetime.strptime(desde_s, "%Y-%m-%d").date()
        dt_desde = timezone.make_aware(datetime.combine(d, time.min), tz)

    if hasta_s:
        d = datetime.strptime(hasta_s, "%Y-%m-%d").date()
        dt_hasta = timezone.make_aware(datetime.combine(d, time.max), tz)

    return dt_desde, dt_hasta


def _base_servicios_qs(request):
    qs = Servicio.objects.select_related("cliente").all()

    dt_desde, dt_hasta = _parse_date_range(request)
    if dt_desde:
        qs = qs.filter(fecha_ingreso__gte=dt_desde)
    if dt_hasta:
        qs = qs.filter(fecha_ingreso__lte=dt_hasta)

    estado = (request.GET.get("estado") or "").strip()
    if estado:
        qs = qs.filter(estado_actual=estado)

    return qs, dt_desde, dt_hasta, estado





def _build_report_data(request):
    # ✅ ahora vienen 4 valores (incluye estado)
    qs, dt_desde, dt_hasta, estado = _base_servicios_qs(request)

    total_servicios = qs.count()
    total_ingresos = qs.aggregate(total=Sum("costo_total"))["total"] or Decimal("0.00")

    por_estado = list(
        qs.values("estado_actual")
          .annotate(cant=Count("id"))
          .order_by("-cant")
    )

    top_clientes = list(
        qs.values("cliente__apellido", "cliente__nombre")
          .annotate(cant=Count("id"))
          .order_by("-cant", "cliente__apellido")[:10]
    )

    detalle = list(
        qs.order_by("-fecha_ingreso").values(
            "id",
            "fecha_ingreso",
            "cliente__apellido",
            "cliente__nombre",
            "tipo_equipo",
            "marca_modelo",
            "estado_actual",
            "costo_total",
        )
    )

    return {
        "dt_desde": dt_desde,
        "dt_hasta": dt_hasta,
        "estado_actual": estado,  # ✅ filtro aplicado (o "" si es "Todos")
        "total_servicios": total_servicios,
        "total_ingresos": float(total_ingresos),
        "por_estado": por_estado,
        "top_clientes": top_clientes,
        "detalle": detalle,
    }

def reportes_export_excel(request):
    data = _build_report_data(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    def header(cell):
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    # Resumen
    ws["A1"] = "Reporte Servicio Técnico"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Desde"; header(ws["A3"])
    ws["B3"] = data["dt_desde"].strftime("%d/%m/%Y") if data["dt_desde"] else "—"

    ws["A4"] = "Hasta"; header(ws["A4"])
    ws["B4"] = data["dt_hasta"].strftime("%d/%m/%Y") if data["dt_hasta"] else "—"

    ws["A6"] = "Total servicios"; header(ws["A6"])
    ws["B6"] = data["total_servicios"]

    ws["A7"] = "Total ingresos"; header(ws["A7"])
    ws["B7"] = data["total_ingresos"]
    ws["B7"].number_format = '"$" #,##0.00'

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22

    # Por estado
    ws2 = wb.create_sheet("Por estado")
    ws2.append(["Estado", "Cantidad"])
    ws2["A1"].font = Font(bold=True); ws2["B1"].font = Font(bold=True)
    for r in data["por_estado"]:
        ws2.append([r["estado_actual"], r["cant"]])
    ws2.freeze_panes = "A2"
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 10

    # Top clientes
    ws3 = wb.create_sheet("Top clientes")
    ws3.append(["Cliente", "Servicios"])
    ws3["A1"].font = Font(bold=True); ws3["B1"].font = Font(bold=True)
    for r in data["top_clientes"]:
        nombre = f'{r["cliente__apellido"]}, {r["cliente__nombre"]}'
        ws3.append([nombre, r["cant"]])
    ws3.freeze_panes = "A2"
    ws3.column_dimensions["A"].width = 32
    ws3.column_dimensions["B"].width = 10

    # Detalle
    ws4 = wb.create_sheet("Servicios")
    ws4.append(["ID", "Ingreso", "Cliente", "Equipo", "Marca/Modelo", "Estado", "Costo"])
    for col in range(1, 8):
        ws4.cell(row=1, column=col).font = Font(bold=True)

    for r in data["detalle"]:
        cliente = f'{r["cliente__apellido"]}, {r["cliente__nombre"]}'
        ingreso = r["fecha_ingreso"].strftime("%d/%m/%Y %H:%M") if r["fecha_ingreso"] else ""
        ws4.append([
            r["id"],
            ingreso,
            cliente,
            r["tipo_equipo"],
            r["marca_modelo"] or "",
            r["estado_actual"],
            float(r["costo_total"] or 0),
        ])

    ws4.freeze_panes = "A2"
    widths = [8, 18, 28, 18, 22, 20, 12]
    for i, w in enumerate(widths, start=1):
        ws4.column_dimensions[get_column_letter(i)].width = w
    for row in range(2, ws4.max_row + 1):
        ws4.cell(row=row, column=7).number_format = '"$" #,##0.00'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"reporte_servicio_tecnico_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp



def reportes_export_pdf(request):
    data = _build_report_data(request)

    # --- Datos del negocio (formal) ---
    EMPRESA = "Servicio Técnico"
    TITULAR = "Victor Soria"
    DIRECCION = "9 de Julio 1070, Metán, Salta"
    TEL = "Cel. 3876505131"

    # --- Helpers ---
    def fmt_fecha(dt):
        return dt.strftime("%d/%m/%Y") if dt else "—"

    def fmt_datetime(dt):
        return dt.strftime("%d/%m/%Y %H:%M") if dt else "—"

    def fmt_ars(value):
        try:
            s = f"{float(value):,.2f}"
        except Exception:
            s = "0.00"
        return s.replace(",", "X").replace(".", ",").replace("X", ".")

    def safe_text(s):
        return (s or "").strip()

    # Filtros mostrados
    estado_str = (request.GET.get("estado") or "").strip() or "Todos"
    rango_desde = fmt_fecha(data.get("dt_desde"))
    rango_hasta = fmt_fecha(data.get("dt_hasta"))
    emitido = timezone.localtime()

    # --- Documento ---
    bio = BytesIO()
    doc = SimpleDocTemplate(
        bio,
        pagesize=A4,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
        title="Reporte Servicio Técnico",
        author=TITULAR,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="Small", parent=styles["Normal"], fontSize=9, leading=11))
    styles.add(ParagraphStyle(name="Tiny", parent=styles["Normal"], fontSize=8, leading=10))
    styles.add(ParagraphStyle(name="H1", parent=styles["Title"], fontSize=18, leading=22, spaceAfter=6))
    styles.add(ParagraphStyle(name="H2", parent=styles["Heading2"], fontSize=12, leading=14, spaceBefore=12, spaceAfter=6))

    story = []

    # --- Encabezado formal ---
    story.append(Paragraph("Reporte", styles["H1"]))
    story.append(Paragraph(f"<b>{EMPRESA}</b>", styles["Normal"]))
    story.append(Paragraph(f"{TITULAR} · {DIRECCION} · {TEL}", styles["Small"]))
    story.append(Paragraph(f"<b>Emitido:</b> {fmt_datetime(emitido)}", styles["Small"]))
    story.append(Spacer(1, 8))

    # --- Filtros aplicados ---
    story.append(Paragraph("<b>Filtros aplicados</b>", styles["Normal"]))
    story.append(Paragraph(f"Rango: {rango_desde} a {rango_hasta}", styles["Small"]))
    story.append(Paragraph(f"Estado: {estado_str}", styles["Small"]))
    story.append(Spacer(1, 10))

    # --- Resumen ---
    resumen = [
        ["Total servicios", str(data["total_servicios"])],
        ["Total ingresos", f"$ {fmt_ars(data['total_ingresos'])}"],
    ]
    t_res = Table(resumen, colWidths=[220, 280])
    t_res.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.grey),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t_res)

    # --- Por estado ---
    story.append(Paragraph("Servicios por estado", styles["H2"]))
    table_estado = [["Estado", "Cantidad"]]
    for r in data.get("por_estado", []):
        table_estado.append([r["estado_actual"], str(r["cant"])])

    t1 = Table(table_estado, colWidths=[360, 140], repeatRows=1)
    t1.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t1)

    # --- Top clientes ---
    story.append(Paragraph("Top clientes (por cantidad de servicios)", styles["H2"]))
    table_clientes = [["Cliente", "Servicios"]]
    for r in data.get("top_clientes", []):
        cliente = f'{r["cliente__apellido"]}, {r["cliente__nombre"]}'
        table_clientes.append([cliente, str(r["cant"])])

    t2 = Table(table_clientes, colWidths=[360, 140], repeatRows=1)
    t2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t2)

    # =========================================================
    # ✅ DETALLE DE SERVICIOS (máx 50)
    # =========================================================
    DETALLE_MAX = 50
    detalle = data.get("detalle", [])[:DETALLE_MAX]

    story.append(PageBreak())
    story.append(Paragraph(f"Detalle de servicios (máx. {DETALLE_MAX})", styles["H2"]))
    story.append(Paragraph("Ordenado por fecha de ingreso (más recientes primero).", styles["Small"]))
    story.append(Spacer(1, 6))

    detalle_table = [["ID", "Ingreso", "Cliente", "Equipo", "Estado", "Costo"]]

    for r in detalle:
        cliente = f'{r["cliente__apellido"]}, {r["cliente__nombre"]}'
        ingreso = r["fecha_ingreso"].strftime("%d/%m/%Y %H:%M") if r["fecha_ingreso"] else ""
        equipo = safe_text(r["tipo_equipo"])
        mm = safe_text(r.get("marca_modelo"))
        if mm:
            equipo = f"{equipo} / {mm}"
        detalle_table.append([
            str(r["id"]),
            ingreso,
            cliente,
            equipo,
            r["estado_actual"],
            f"$ {fmt_ars(r.get('costo_total') or 0)}",

        ])

    t3 = Table(
        detalle_table,
        colWidths=[45, 85, 130, 150, 70, 70],
        repeatRows=1
    )
    t3.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 1), (0, -1), "RIGHT"),   # ID
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),   # Costo
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(t3)

    # --- Footer con numeración ---
    def _on_page(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 9)
        canvas.setFillGray(0.35)
        canvas.drawRightString(A4[0] - 36, 20, f"Página {doc_.page}")
        canvas.drawString(36, 20, f"{EMPRESA} · {TITULAR}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    bio.seek(0)

    filename = f"reporte_servicio_tecnico_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
    resp = HttpResponse(bio.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

# =========================================================
    # ✅ COMPROBANTE DE SERVICIOS
    # =========================================================
class ServicioComprobanteView(DetailView):
    model = Servicio
    template_name = "core/servicio_comprobante.html"